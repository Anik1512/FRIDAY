import openpyxl
import os

# Function to load or create an Excel file for a subject
def load_or_create_file(subject):
    filename = f"{subject}.xlsx"
    if not os.path.exists(filename):
        # Create a new workbook with headers if file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = subject
        sheet.append(["Roll", "Name"])  # Initial headers
        workbook.save(filename)
    return filename

# Function to update attendance
def update_attendance(subject, date, present_rolls):
    filename = load_or_create_file(subject)
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Check if date column exists, if not add it
    if date not in [cell.value for cell in sheet[1]]:
        sheet.cell(row=1, column=sheet.max_column + 1, value=date)

    # Map the date to its column number
    date_column = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == date:
            date_column = col
            break

    # Create a set of rolls for quick lookup
    roll_name_mapping = {sheet.cell(row=row, column=1).value: sheet.cell(row=row, column=2).value for row in range(2, sheet.max_row + 1)}

    print("\nMarked Present:")
    for roll in present_rolls:
        # Print only the inputed roll numbers with names (if available)
        if int(roll) in roll_name_mapping:
            print(f"Roll: {roll}, Name: {roll_name_mapping[int(roll)]}")
        else:
            print(f"Roll: {roll} not found in the file.")

    # Update attendance for each roll number in the sheet
    for row in range(2, sheet.max_row + 1):
        roll = sheet.cell(row=row, column=1).value
        if roll is not None and str(roll) in present_rolls:
            sheet.cell(row=row, column=date_column, value=1)  # Mark as present
        else:
            sheet.cell(row=row, column=date_column, value=0)  # Mark as absent

    workbook.save(filename)
    print(f"Attendance updated for {subject} on {date}.\n")

# Function to handle Event-specific attendance
def manage_event(date, present_rolls):
    filename = "Event.xlsx"
    workbook = openpyxl.Workbook() if not os.path.exists(filename) else openpyxl.load_workbook(filename)
    sheet = workbook.active

    # Add headers if the file is newly created
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value != "Roll":
        sheet.append(["Roll", "Name", date])
    elif date not in [cell.value for cell in sheet[1]]:
        # Add date column if it doesn't exist
        sheet.cell(row=1, column=sheet.max_column + 1, value=date)

    # Map the date to its column number
    date_column = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == date:
            date_column = col
            break

    print("\nMarked Present for Event:")
    for roll in present_rolls:
        row_found = False
        # Check if roll already exists
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == roll:
                row_found = True
                sheet.cell(row=row, column=date_column, value=1)  # Mark as present
                print(f"Roll: {roll}, Name: {sheet.cell(row=row, column=2).value}")
                break

        # If roll not found, append it
        if not row_found:
            sheet.append([roll, "", 1 if date_column == 3 else None])
            print(f"Roll: {roll} added to the file.")

    workbook.save(filename)
    print(f"Attendance updated for Event on {date}.\n")

# Main function to interact with the user
def main():
    # List of subjects
    subjects = ["Multimedia", "CHN", "MC", "IOT", "Java", "Event"]

    # Ask for date
    date = input("Enter the date (DD-MM-YYYY): ")

    # Choose subject
    print("Choose a subject:")
    for i, subject in enumerate(subjects, start=1):
        print(f"{i}. {subject}")
    subject_choice = int(input("Enter the number corresponding to the subject: "))
    subject = subjects[subject_choice - 1]

    # Get present rolls
    roll_list = input("Enter the roll numbers of present students, separated by commas: ")
    present_rolls = set(roll.strip() for roll in roll_list.split(","))

    # Handle Event separately
    if subject == "Event":
        manage_event(date, present_rolls)
    else:
        update_attendance(subject, date, present_rolls)


